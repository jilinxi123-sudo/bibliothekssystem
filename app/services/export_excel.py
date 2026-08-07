from io import BytesIO
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models import BookOut
from app.services.export_columns import (
    column_labels,
    history_text,
    is_history_column,
    is_image_column,
    text_value,
)

COLUMN_WIDTHS = {
    "isbn": 16, "title": 30, "author": 24, "publisher": 20, "published_year": 8,
    "themes": 20, "source": 16, "location": 16, "notes": 30, "due_date": 14,
    "created_at": 20, "updated_at": 20, "cover": 10, "history": 45,
}

COVER_IMG_WIDTH_PX = 45
COVER_IMG_HEIGHT_PX = 60
COVER_ROW_HEIGHT_PT = 48


def build_excel_export(
    books: list[BookOut],
    columns: list[str],
    covers_dir: Optional[Path] = None,
    history_by_isbn: Optional[dict[str, list[dict]]] = None,
) -> bytes:
    history_by_isbn = history_by_isbn or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Bibliothek"

    ws.append(column_labels(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for book in books:
        row_values = []
        for key in columns:
            if is_image_column(key):
                row_values.append("")
            elif is_history_column(key):
                row_values.append(history_text(history_by_isbn.get(book.isbn, [])))
            else:
                row_values.append(text_value(book, key))
        ws.append(row_values)
        row_num = ws.max_row

        for i, key in enumerate(columns):
            col_letter = get_column_letter(i + 1)
            if is_history_column(key):
                ws.cell(row=row_num, column=i + 1).alignment = Alignment(wrap_text=True, vertical="top")
            if is_image_column(key) and covers_dir is not None:
                cover_path = covers_dir / f"{book.isbn}.jpg"
                if cover_path.exists():
                    img = XLImage(str(cover_path))
                    img.width = COVER_IMG_WIDTH_PX
                    img.height = COVER_IMG_HEIGHT_PX
                    img.anchor = f"{col_letter}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = COVER_ROW_HEIGHT_PT

    for i, key in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(key, 20)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
